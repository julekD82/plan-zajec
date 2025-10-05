import requests
from bs4 import BeautifulSoup
import os
import pandas as pd
from sqlalchemy import create_engine
from datetime import datetime
import re
import locale
from openpyxl import load_workbook
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
import sys

# Konfiguracja dla Render.com
if 'RENDER' in os.environ:
    # Render.com specific configuration
    DB_URL = 'sqlite:///plan.db'
    FILE_PATH = '/tmp/plan_downloaded.xlsx'
else:
    # Local development
    DB_URL = 'sqlite:///plan.db'
    FILE_PATH = 'plan_downloaded.xlsx'

try:
    locale.setlocale(locale.LC_TIME, 'pl_PL.UTF-8')
except locale.Error:
    print("Nie można ustawić polskich locale. Dni tygodnia mogą być po angielsku.")

TABLE_NAME = 'schedule_entries'
CHECK_URL = 'https://www.ur.edu.pl/pl/collegium-medicum-2/collegium-medicum/jednostki/wydzial-medyczny/student/lekarski/rozklady-zajec'

def determine_gradient_class(subject):
    """Determines the CSS gradient class based on the subject name."""
    subject_lower = subject.lower()
    if " w1" in subject_lower or " w2" in subject_lower or "wykład" in subject_lower:
        return "accent-blue-gradient"
    elif "sem" in subject_lower or "seminarium" in subject_lower:
        return "accent-cyan-gradient"
    elif "sala" in subject_lower or "ćwiczenia" in subject_lower:
        return "accent-green-gradient"
    else:
        return "accent-orange-gradient"

def parse_time_range(text):
    """
    Ulepszona funkcja do parsowania zakresu czasu.
    Wyszukuje dwa pierwsze wystąpienia formatów czasu jak "8", "8.00", "13:15".
    Zwraca (start_time, end_time) w formacie HH:MM lub (None, None).
    """
    if not isinstance(text, str):
        return None, None
    
    # Ten wzorzec (wyrażenie regularne) znajduje liczby, które mogą być godzinami,
    # z opcjonalną częścią minutową (np. '8', '8.00', '13:15').
    pattern = re.compile(r'\b(\d{1,2}(?:[.:]\d{2})?)\b')
    matches = pattern.findall(text)
    
    if len(matches) < 2:
        return None, None # Nie znaleziono wystarczającej liczby znaczników czasu

    def normalize_time(time_str):
        """Pomocnicza funkcja do konwersji '8' na '08:00' lub '13.15' na '13:15'."""
        time_str = time_str.replace('.', ':')
        if ':' not in time_str:
            return f"{int(time_str):02d}:00"
        else:
            parts = time_str.split(':')
            return f"{int(parts[0]):02d}:{int(parts[1]):02d}"

    try:
        # Bierzemy dwie pierwsze znalezione liczby jako czas startu i końca
        start_time = normalize_time(matches[0])
        end_time = normalize_time(matches[1])
        return start_time, end_time
    except (ValueError, IndexError):
        # Zabezpieczenie na wypadek, gdyby znalezione liczby nie były poprawnym czasem
        return None, None

def get_cell_color(cell, sheet):
    """
    Ulepszona funkcja do pobierania koloru tła komórki w formacie HEX (#RRGGBB).
    Poprawnie obsługuje komórki scalone i kolory z motywu Excela.
    """
    # Krok 1: Jeśli komórka jest scalona, znajdź jej komórkę główną (lewy górny róg),
    # ponieważ styl, w tym kolor, jest zapisany tylko w tej jednej komórce.
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
            break # Znaleziono, można przerwać pętlę

    # Krok 2: Pobierz informacje o wypełnieniu komórki.
    # W przypadku jednolitego tła kolor znajduje się w atrybucie `fgColor` (kolor pierwszoplanowy).
    fill = cell.fill
    
    # Krok 3: Wyodrębnij właściwy obiekt koloru.
    color_obj = None
    if fill.patternType == 'solid':
        color_obj = fill.fgColor
    elif fill.patternType:
         # Dla innych wzorów (np. gradientów) `start_color` może być istotny.
        color_obj = fill.start_color

    # Krok 4: Przekonwertuj obiekt koloru na format HEX (#RRGGBB).
    if color_obj and hasattr(color_obj, 'rgb') and color_obj.rgb:
        # Wartość `rgb` jest zazwyczaj w formacie AARRGGBB (Alfa, Czerwony, Zielony, Niebieski).
        # My potrzebujemy tylko ostatnich 6 znaków, czyli RRGGBB.
        rgb_hex = color_obj.rgb[-6:]
        
        # Czasami "brak koloru" jest błędnie interpretowany jako czarny ('000000').
        # Na wszelki wypadek ignorujemy go, aby nie kolorować pustych komórek.
        if rgb_hex != '000000':
            return f"#{rgb_hex}"
            
    # Jeśli nie udało się znaleźć żadnego koloru, zwróć None.
    return None



def load_and_process_data_rok6(file_path):
    """
    Loads and processes the new schedule format for the 6th year.
    The schedule is laid out horizontally (days in columns).
    """

    from openpyxl import load_workbook
    wb = load_workbook(file_path, data_only=True)
    sheet = wb["semestr 11"]
    try:
        df = pd.read_excel(file_path, sheet_name='semestr 11', header=None)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return []

    all_entries = []

    GROUP_COL_IDX = 18
    first_group_row = -1
    for i, val in enumerate(df[GROUP_COL_IDX]):
        if isinstance(val, (int, float)) and not pd.isna(val):
            first_group_row = i
            break

    if first_group_row == -1:
        print("Could not find the starting row for groups in column S.")
        return []

    date_row_idx = 3
    date_row = df.iloc[date_row_idx]
    first_date_col = -1
    for i in range(GROUP_COL_IDX + 1, len(date_row)):
        if isinstance(date_row[i], datetime):
            first_date_col = i
            break

    if first_date_col == -1:
        print("Could not find the starting column for dates.")
        return []
    #  --- Inteligentne wypełnianie na podstawie scalonych komórek z pliku XLSX ---

    # Iterujemy po wszystkich scalonych zakresach w arkuszu. To jest klucz do poprawnego odczytu struktury.
    for merged_range in sheet.merged_cells.ranges:
        # Pobieramy współrzędne (granice) każdego scalonego bloku
        min_col, min_row, max_col, max_row = merged_range.bounds
        
        # Odczytujemy wartość z lewej górnej komórki scalonego bloku
        top_left_cell_value = sheet.cell(row=min_row, column=min_col).value

        # Jeśli komórka nie jest pusta, to jej wartość jest "rozsiewana" na cały obszar,
        # który obejmował scalony blok w DataFrame.
        if top_left_cell_value:
            # Używamy .loc do przypisania wartości. Pamiętamy o konwersji indeksów:
            # openpyxl jest 1-based (liczy od 1), a pandas jest 0-based (liczy od 0).
            # Dlatego odejmujemy 1 od każdego indeksu wiersza i kolumny.
            df.loc[min_row-1:max_row-1, min_col-1:max_col-1] = top_left_cell_value


    col_to_date_map = {}
    current_date = None
    for i in range(first_date_col, len(date_row)):
        if isinstance(date_row[i], datetime):
            current_date = date_row[i]
        if current_date:
            col_to_date_map[i] = current_date

    last_end_by_date_group = {}

    for index, row in df.iloc[first_group_row:].iterrows():
        group_number = row[GROUP_COL_IDX]

        if pd.isna(group_number) or not str(group_number).strip():
            continue

        try:
            group_number = int(group_number)
        except (ValueError, TypeError):
            continue

        for col_idx in range(first_date_col, len(row)):
            cell_value = row[col_idx]

            if pd.notna(cell_value) and str(cell_value).strip():
                # Konwertujemy `cell_value` na string, aby reszta kodu w tym bloku
                # (np. parse_time_range) działała poprawnie.
                cell_value = str(cell_value)
                current_date = col_to_date_map.get(col_idx)
                if not current_date:
                    continue

                start_time_str, end_time_str = parse_time_range(cell_value)

                if start_time_str and end_time_str:
                    day_start_minutes = 7 * 60 -30               
                    key = (current_date.date(), group_number)
                    last_end = last_end_by_date_group.get(key, day_start_minutes)

                   
                   
                    # Przetwarzanie czasu na minuty
                    start_h, start_m = map(int, start_time_str.split(':'))
                    end_h, end_m = map(int, end_time_str.split(':'))
                    start_time_total_minutes = start_h * 60 + start_m
                    end_time_total_minutes = end_h * 60 + end_m

                    spacing_before = max(0, start_time_total_minutes - last_end)


                    duration = end_time_total_minutes - start_time_total_minutes
                    if duration < 0: continue # Sanity check
                    # Obliczanie pustych slotów przed zajęciami (zakładając dzień od 7:30)
                    

                    subject = ' '.join(cell_value.replace('\n', ' ').split())
                    POLISH_DAYS = {
                        0: "PONIEDZIAŁEK",
                        1: "WTOREK",
                        2: "ŚRODA",
                        3: "CZWARTEK",
                        4: "PIĄTEK",
                        5: "Sobota",
                        6: "Niedziela"
                    }

                    day_name = POLISH_DAYS[current_date.weekday()]

                    # Otwórz skoroszyt (tylko raz, poza pętlą – dodamy to niżej)
                    # workbook = load_workbook(file_path, data_only=True)
                    # sheet = workbook["semestr 11"]

                    # Pobierz kolor komórki (zakładamy, że masz dostęp do ws[row][col])
                    excel_row = index + 1  # openpyxl używa indeksowania od 1
                    excel_col_letter = get_column_letter(col_idx + 1)
                    cell_ref = f"{excel_col_letter}{excel_row}"

                    color = None
                    try:
                        # sheet musi być zdefiniowany przed pętlą – dodamy to niżej
                        cell = sheet[cell_ref]
                        color = get_cell_color(cell, sheet)
                    except Exception:
                        color = None

                    entry = {
                        'date': current_date.date(),
                        'day': day_name,
                        'group_number': str(group_number),
                        'subject': subject,
                        'start_time_formatted': f"{start_time_total_minutes//60:02d}:{start_time_total_minutes%60:02d}",
                        'end_time_formatted': f"{(start_time_total_minutes+duration)//60:02d}:{(start_time_total_minutes+duration)%60:02d}",
                        'duration': duration,
                        'spacing_before': spacing_before,
                        'background_color': color or "#FFD966"  # domyślny kolor, jeśli brak
                    }

                    all_entries.append(entry)

                    last_end_by_date_group[key] = end_time_total_minutes

    return all_entries


def get_last_update_info():
    """Reads the last update date and URL from the log file."""
    try:
        with open('last_update.txt', 'r') as f:
            line = f.readline().strip()
            if '|' in line:
                return line.split('|')
            return line, ""
    except FileNotFoundError:
        return None, None

def save_last_update_info(date_str, url):
    """Saves the new update date and URL to the log file."""
    with open('last_update.txt', 'w') as f:
        f.write(f"{date_str}|{url}")

def main():
    """Main function to check for updates, download, process, and save schedule data."""
    try:
        response = requests.get(CHECK_URL)
        response.raise_for_status()
    except requests.RequestException as e:
        print(f"Error fetching the website: {e}")
        return

    soup = BeautifulSoup(response.text, 'html.parser')
    
    # Find the link for the 6th year schedule
    schedule_link_tag = soup.find('a', href=lambda h: h and 'VI%20rok' in h)

    if not schedule_link_tag:
        print("Could not find the schedule link for the 6th year.")
        return

    link = schedule_link_tag['href']

    # The update text is usually within a span or the link itself. Let's find the closest description.
    update_text = schedule_link_tag.find_parent('tr').get_text(strip=True)
    date_match = re.search(r'(\d{1,2}\.\d{1,2}\.\d{4})', update_text)




    if date_match:
        update_date_str = date_match.group(1)

    if not update_date_str:
        print("Could not parse update date from website.")
        # Fallback to current date to ensure it runs at least once
        update_date_str = datetime.now().strftime('%d.%m.%Y')

    last_date, last_url = get_last_update_info()

    if last_date != update_date_str or last_url != link:
        print("New schedule update detected. Downloading and processing...")
        
        
        download_url = 'https://www.ur.edu.pl/' + link
        

        print(f"Downloading from: {download_url}")
        
        try:
            file_response = requests.get(download_url)
            file_response.raise_for_status()
            with open(FILE_PATH, 'wb') as f:
                f.write(file_response.content)
        except requests.RequestException as e:
            print(f"Error downloading the file: {e}")
            return

        # Use the new processing function
        processed_data = load_and_process_data_rok6(FILE_PATH)

        if not processed_data:
            print("No data processed. Aborting database update.")
            return

        df_to_save = pd.DataFrame(processed_data)
        
        try:
            engine = create_engine(DB_URL)
            df_to_save.to_sql(TABLE_NAME, engine, if_exists='replace', index=False)
            save_last_update_info(update_date_str, link)
            print(f"Successfully processed and saved {len(df_to_save)} entries to the database.")
        except Exception as e:
            print(f"Error saving data to database: {e}")
            
    else:
        print("Schedule is up to date. No changes were made.")

if __name__ == '__main__':
    main()
