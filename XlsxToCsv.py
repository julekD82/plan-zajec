import pandas as pd
from openpyxl import load_workbook

# Wczytaj plik Excel
file_path = 'plan30.xlsx'  # Zmień na ścieżkę do swojego pliku
workbook = load_workbook(file_path)
sheet = workbook.active  # Wybierz aktywny arkusz

# Przygotuj listę na dane
data = []

# Iteruj przez wszystkie wiersze w arkuszu
for row in sheet.iter_rows(values_only=True):
    new_row = []
    for cell in row:
        if cell is None:
            new_row.append('')  # Zastąp puste komórki pustym stringiem
        else:
            new_row.append(cell)  # Dodaj wartość komórki
    data.append(new_row)

# Stwórz DataFrame z danych
df = pd.DataFrame(data)

# Sprawdź liczbę wierszy
print(f'Liczba wierszy w DataFrame: {len(df)}')

# Zastąp puste wartości w scalonych komórkach
for row in range(len(df)):
    for col in range(len(df.columns)):
        if df.iloc[row, col] == '':
            # Sprawdź, czy ta komórka jest w scalonym zakresie
            merged_cells = sheet.merged_cells.ranges  # Zbieranie scalonych komórek
            for merged_range in merged_cells:
                min_col, min_row, max_col, max_row = merged_range.bounds
                # Sprawdzenie, czy komórka jest w scalonym zakresie
                if (col + 1 >= min_col and col + 1 <= max_col and 
                    row + 1 >= min_row and row + 1 <= max_row):
                    # Jeśli jest scalona, przypisz wartość z pierwszej komórki
                    df.iloc[row, col] = sheet.cell(row=min_row, column=min_col).value

# Zapisz do CSV
csv_file_path = 'scalony_plik.csv'  # Ścieżka do zapisanego pliku CSV
df.to_csv(csv_file_path, index=False)

print(f'Dane zostały zapisane do {csv_file_path}')
